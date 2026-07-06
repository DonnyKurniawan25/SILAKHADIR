"""Export laporan kinerja harian ke format Excel."""
import io
from datetime import datetime

from django.http import HttpResponse

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


BULAN_INDO = {
    1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
    5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
    9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember',
}


def export_laporan_excel(periode, entries):
    """Generate Excel file for laporan kinerja."""
    if not HAS_OPENPYXL:
        return HttpResponse(
            'Library openpyxl belum terinstal. Jalankan: pip install openpyxl',
            status=500,
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laporan Kinerja'

    # Styles
    header_font = Font(name='Calibri', size=14, bold=True)
    sub_font = Font(name='Calibri', size=11, bold=True)
    col_header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    cell_font = Font(name='Calibri', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color='0C4469', end_color='0C4469', fill_type='solid')

    # Title
    bulan_nama = BULAN_INDO.get(periode.bulan, str(periode.bulan))
    ws.merge_cells('A1:I1')
    ws['A1'] = f'LAPORAN KINERJA HARIAN ASN'
    ws['A1'].font = header_font
    ws['A1'].alignment = center

    ws.merge_cells('A2:I2')
    ws['A2'] = f'{periode.nama}'
    ws['A2'].font = sub_font
    ws['A2'].alignment = center

    ws.merge_cells('A3:I3')
    ws['A3'] = f'Bulan {bulan_nama} {periode.tahun} — Bidang {periode.bidang}'
    ws['A3'].font = Font(name='Calibri', size=10)
    ws['A3'].alignment = center

    # Column headers
    row = 5
    headers = ['No', 'Tanggal', 'NIP', 'Nama Pegawai', 'Email', 'No HP', 'Jabatan', 'Uraian Kegiatan',
               'Link Bukti']
    widths = [5, 14, 22, 25, 25, 18, 25, 45, 35]

    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = col_header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # Data rows
    row = 6
    for idx, entry in enumerate(entries, 1):
        nama_pegawai = entry.pegawai.get_full_name() or entry.pegawai.username if entry.pegawai else entry.nama_pegawai
        nip_pegawai = entry.pegawai.nip if entry.pegawai and entry.pegawai.nip else entry.nip_pegawai
        email_pegawai = entry.pegawai.email if entry.pegawai and entry.pegawai.email else entry.email_pegawai
        no_hp_pegawai = entry.pegawai.phone if entry.pegawai and entry.pegawai.phone else entry.no_hp_pegawai
        jabatan_pegawai = entry.pegawai.jabatan if entry.pegawai and entry.pegawai.jabatan else entry.jabatan_pegawai

        data = [
            idx,
            entry.tanggal.strftime('%d/%m/%Y'),
            nip_pegawai or '-',
            nama_pegawai,
            email_pegawai or '-',
            no_hp_pegawai or '-',
            jabatan_pegawai or '-',
            entry.uraian_kegiatan,
            entry.link_bukti or '',
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = left if col_idx >= 7 else center
            cell.border = thin_border
        row += 1

    # Footer
    row += 1
    ws.merge_cells(f'A{row}:I{row}')
    ws.cell(row=row, column=1, value=f'Total entri: {entries.count()}').font = sub_font
    row += 1
    ws.cell(
        row=row, column=1,
        value=f'Dicetak pada: {datetime.now().strftime("%d/%m/%Y %H:%M")}',
    ).font = Font(name='Calibri', size=9, italic=True)

    # Response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'Laporan_Kinerja_{periode.bidang}_{bulan_nama}_{periode.tahun}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
