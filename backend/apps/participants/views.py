import io

import openpyxl
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

from apps.accounts.permissions import IsAuthenticatedStaff
from apps.events.models import Event

from .models import Participant
from .serializers import ParticipantSerializer


class EventParticipantViewSet(viewsets.ModelViewSet):
    serializer_class = ParticipantSerializer
    permission_classes = [IsAuthenticatedStaff]
    search_fields = ['full_name', 'nik', 'nip', 'institution', 'email']
    filterset_fields = ['is_asn']

    def get_queryset(self):
        event_id = self.kwargs.get('event_id')
        return (
            Participant.objects
            .filter(event_id=event_id)
            .select_related('event')
            .prefetch_related('attendances', 'certificates')
        )

    def _get_event(self):
        return get_object_or_404(Event, id=self.kwargs['event_id'])

    def perform_create(self, serializer):
        serializer.save(event=self._get_event())

    @action(detail=False, methods=['post'], url_path='import-excel',
            parser_classes=[MultiPartParser, FormParser])
    def import_excel(self, request, event_id=None):
        event = self._get_event()
        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'File Excel wajib diunggah.'},
                            status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            headers = [str(c.value or '').strip().lower() for c in ws[1]]
            required = {'nik', 'full_name'}
            if not required.issubset(set(headers)):
                return Response({
                    'detail': f'Header wajib: {", ".join(sorted(required))}',
                }, status=status.HTTP_400_BAD_REQUEST)

            created, skipped, errors = 0, 0, []
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                data = dict(zip(headers, row))
                nik = str(data.get('nik') or '').strip()
                if not nik:
                    continue
                exists = Participant.objects.filter(
                    event=event, nik=nik
                ).exists()
                if exists:
                    skipped += 1
                    continue
                nip = str(data.get('nip') or '').strip()
                is_asn = str(data.get('is_asn') or '').strip().lower() in ('true', '1', 'yes')
                if not is_asn and nip:
                    is_asn = True
                try:
                    Participant.objects.create(
                        event=event,
                        nik=nik,
                        nip=nip,
                        is_asn=is_asn,
                        full_name=data.get('full_name') or '',
                        institution=data.get('institution') or '',
                        position=data.get('position') or '',
                        phone=str(data.get('phone') or ''),
                        email=data.get('email') or '',
                    )
                    created += 1
                except Exception as exc:
                    errors.append({'row': i, 'error': str(exc)})
            return Response({
                'created': created, 'skipped': skipped, 'errors': errors,
            })
        except Exception as exc:
            return Response({'detail': f'Gagal membaca file: {exc}'},
                            status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request, event_id=None):
        event = self._get_event()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Peserta'
        ws.append([
            'nik', 'nip', 'is_asn', 'full_name', 'institution',
            'position', 'phone', 'email', 'attendance_status', 'attendance_time',
        ])
        for p in self.get_queryset():
            att = p.attendances.first()
            ws.append([
                p.nik,
                p.nip,
                p.is_asn,
                p.full_name,
                p.institution,
                p.position,
                p.phone,
                p.email,
                att.status if att else 'belum_hadir',
                att.attendance_time.strftime('%Y-%m-%d %H:%M:%S') if att else '',
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f'peserta-{event.public_slug}.xlsx'
        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'
        return resp


class ParticipantViewSet(viewsets.ModelViewSet):
    """Endpoint flat /api/participants/{id}/ untuk edit/hapus satu peserta."""
    queryset = Participant.objects.select_related('event')
    serializer_class = ParticipantSerializer
    permission_classes = [IsAuthenticatedStaff]
